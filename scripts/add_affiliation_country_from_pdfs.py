#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Extract the country of the FIRST affiliation of the FIRST author from PDFs and
add it to an OSF-style workbook (January..December sheets).

Workflow:
- Recursively scan PDFs under --year-folder
- Extract DOI from first pages -> build doi -> pdf_path index
- For each row in the workbook (by DOI in Link):
  - open matching PDF
  - extract first-page text
  - locate affiliation line (heuristic, before Abstract)
  - infer country from affiliation
  - write country to new column "First author affiliation country"

Outputs:
- Updates workbook in place (creates timestamped backup by default)
- Writes log CSV: pdf_affiliation_country_log.csv

Install:
  pip install openpyxl pycountry pdfminer.six
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import shutil
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import pycountry
from openpyxl import load_workbook

# Prefer pdfminer for better text extraction than PyPDF2
from pdfminer.high_level import extract_text as pdfminer_extract_text


MONTH_SHEETS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December","Sheet7"
]

NEW_COL = "First author affiliation country"

DOI_RE = re.compile(r"\b10\.\d{4,9}/[^\s<>\"']+\b", re.IGNORECASE)

# Heuristic: stop scanning after "Abstract" on the first page
ABSTRACT_RE = re.compile(r"(?im)^\s*abstract\s*$")

# Affiliations often start with 1, 2, a, b, etc.
AFFIL_LEAD_RE = re.compile(r"^\s*(\d+|[a-z])[\)\.\:]?\s+(.*\S)\s*$", re.IGNORECASE)

# Country aliases that pycountry sometimes misses in substring matching
ALIASES = {
    "usa": "United States",
    "u.s.a.": "United States",
    "u.s.": "United States",
    "united states of america": "United States",
    "uk": "United Kingdom",
    "u.k.": "United Kingdom",
    "england": "United Kingdom",
    "scotland": "United Kingdom",
    "wales": "United Kingdom",
    "russia": "Russian Federation",
    "south korea": "Korea, Republic of",
    "north korea": "Korea, Democratic People's Republic of",
    "iran": "Iran, Islamic Republic of",
    "tanzania": "Tanzania, United Republic of",
    "viet nam": "Viet Nam",
    "czech republic": "Czechia",
    "bolivia": "Bolivia, Plurinational State of",
    "venezuela": "Venezuela, Bolivarian Republic of",
    "moldova": "Moldova, Republic of",
    "laos": "Lao People's Democratic Republic",
    "syria": "Syrian Arab Republic",
}


def build_country_matchers() -> Tuple[Dict[str, str], List[str]]:
    name_map: Dict[str, str] = {}
    for c in pycountry.countries:
        name_map[c.name.lower()] = c.name
        if hasattr(c, "official_name"):
            name_map[c.official_name.lower()] = c.name
        if hasattr(c, "common_name"):
            name_map[c.common_name.lower()] = c.name
    for k, v in ALIASES.items():
        name_map[k.lower()] = v
    keys_sorted = sorted(name_map.keys(), key=len, reverse=True)
    return name_map, keys_sorted


def infer_country(text: str, name_map: Dict[str, str], keys_sorted: List[str]) -> str:
    if not text:
        return ""
    t = " " + text.lower().strip() + " "
    # normalize punctuation/spaces a bit
    t = re.sub(r"[\(\)\[\]\{\};]", " ", t)
    t = re.sub(r"\s+", " ", t)

    for k in keys_sorted:
        if f" {k} " in t:
            return name_map[k]
    return ""


def parse_doi(s: str) -> str:
    if not s:
        return ""
    m = DOI_RE.search(s)
    return m.group(0).lower() if m else ""


def extract_first_pages_text(pdf_path: Path, max_pages: int = 2) -> str:
    """
    Extract text from the first max_pages using pdfminer.six (generally reliable).
    """
    try:
        # pdfminer can limit pages via page_numbers (0-indexed)
        page_numbers = list(range(max_pages))
        return pdfminer_extract_text(str(pdf_path), page_numbers=page_numbers) or ""
    except Exception:
        return ""


def split_pre_abstract(text: str) -> str:
    """
    Keep only text before the "Abstract" heading if found.
    """
    if not text:
        return ""
    lines = text.splitlines()
    out = []
    for ln in lines:
        if ABSTRACT_RE.match(ln.strip()):
            break
        out.append(ln)
    return "\n".join(out)


def pick_first_affiliation_line(pre_abstract_text: str, name_map: Dict[str, str], keys_sorted: List[str]) -> Tuple[str, str]:
    """
    Heuristic to pick first affiliation line (the '1 ...' or 'a ...' line),
    and infer country from it.

    Returns: (affiliation_line, country)
    """
    if not pre_abstract_text:
        return "", ""

    lines = [ln.strip() for ln in pre_abstract_text.splitlines() if ln.strip()]
    # First pass: look for numbered/lettered affiliation lines that contain a country
    for ln in lines:
        m = AFFIL_LEAD_RE.match(ln)
        if not m:
            continue
        body = m.group(2)
        country = infer_country(body, name_map, keys_sorted)
        if country:
            return body, country

    # Second pass: any line that looks affiliation-like and contains country
    for ln in lines:
        # “Department/University/Institute/Hospital” are common affiliation markers
        if re.search(r"(?i)\b(department|university|institute|hospital|centre|center|laboratory|lab)\b", ln):
            country = infer_country(ln, name_map, keys_sorted)
            if country:
                return ln, country

    # Third pass: any line before abstract containing a country
    for ln in lines:
        country = infer_country(ln, name_map, keys_sorted)
        if country:
            return ln, country

    return "", ""


def ensure_column(ws, col_name: str) -> int:
    header = [c.value for c in ws[1]]
    if col_name in header:
        return header.index(col_name) + 1
    ws.cell(row=1, column=len(header) + 1, value=col_name)
    return len(header) + 1


def index_pdfs_by_doi(year_folder: Path, max_pages_for_doi: int = 2) -> Dict[str, Path]:
    """
    Build a mapping DOI -> pdf_path by scanning PDFs and extracting DOI from first pages.
    """
    doi_to_pdf: Dict[str, Path] = {}
    pdfs = sorted([p for p in year_folder.rglob("*.pdf") if p.is_file()])

    for pdf in pdfs:
        text = extract_first_pages_text(pdf, max_pages=max_pages_for_doi)
        doi = parse_doi(text)
        if doi and doi not in doi_to_pdf:
            doi_to_pdf[doi] = pdf

    return doi_to_pdf


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year-folder", required=True, help="Folder containing month folders and PDFs (recursive scan).")
    ap.add_argument("--xlsx", required=True, help="Path to the OSF-style workbook to update.")
    ap.add_argument("--no-backup", action="store_true", help="Do not create a backup XLSX.")
    ap.add_argument("--max-pages-doi", type=int, default=2, help="Pages to scan for DOI in each PDF.")
    ap.add_argument("--max-pages-affil", type=int, default=2, help="Pages to scan for affiliation extraction.")
    args = ap.parse_args()

    year_folder = Path(args.year_folder).expanduser()
    xlsx_path = Path(args.xlsx).expanduser()

    if not year_folder.exists():
        raise SystemExit(f"Year folder not found: {year_folder}")
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    # Backup workbook
    if not args.no_backup:
        backup = xlsx_path.with_suffix(f".backup_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(xlsx_path, backup)
        print(f"Backup written: {backup}")

    # Build country matcher
    name_map, keys_sorted = build_country_matchers()

    # Index PDFs by DOI
    print("Indexing PDFs by DOI (this can take a few minutes)...")
    doi_to_pdf = index_pdfs_by_doi(year_folder, max_pages_for_doi=args.max_pages_doi)
    print(f"Indexed {len(doi_to_pdf)} PDFs with DOIs.")

    # Load workbook
    wb = load_workbook(xlsx_path)

    log_rows = []
    updated = 0
    scanned = 0

    for sheet_name in MONTH_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        if not header or "Link" not in header:
            continue

        link_col = header.index("Link") + 1
        out_col = ensure_column(ws, NEW_COL)

        for r in range(2, ws.max_row + 1):
            link = ws.cell(r, link_col).value
            doi = parse_doi(str(link or ""))
            if not doi:
                continue

            scanned += 1
            pdf_path = doi_to_pdf.get(doi)

            if not pdf_path:
                log_rows.append([sheet_name, r, doi, "", "", "pdf_not_found_for_doi"])
                continue

            text = extract_first_pages_text(pdf_path, max_pages=args.max_pages_affil)
            pre_abs = split_pre_abstract(text)
            affil_line, country = pick_first_affiliation_line(pre_abs, name_map, keys_sorted)

            if country:
                ws.cell(r, out_col).value = country
                updated += 1
                log_rows.append([sheet_name, r, doi, str(pdf_path), country, "ok"])
            else:
                ws.cell(r, out_col).value = ""
                log_rows.append([sheet_name, r, doi, str(pdf_path), "", "country_not_found_in_pdf_text"])

    wb.save(xlsx_path)

    log_path = xlsx_path.parent / "pdf_affiliation_country_log.csv"
    with log_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "row", "doi", "pdf_path", "country", "status"])
        w.writerows(log_rows)

    print(f"Workbook updated: {xlsx_path}")
    print(f"Rows scanned (with DOI): {scanned}")
    print(f"Countries filled: {updated}")
    print(f"Log written: {log_path}")


if __name__ == "__main__":
    main()
