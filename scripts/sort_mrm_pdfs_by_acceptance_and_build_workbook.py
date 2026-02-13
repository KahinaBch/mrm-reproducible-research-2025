#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sort MRM PDFs into month folders based on the *Accepted* date in the PDF,
and build a new OSF-style workbook.

Key improvements vs earlier version:
- Scans PDFs RECURSIVELY (so reruns won't miss PDFs already in month folders).
- Much more robust Accepted-date parsing:
    * normalizes whitespace / non-breaking spaces
    * tolerates line breaks between tokens
    * supports "Accepted: 2 April 2025" and "Accepted 2 April 2025"
    * supports "Accepted: April 2, 2025"
    * supports "Accepted: 2025-04-02" and "Accepted: 02/04/2025" (best-effort)
- Writes a log CSV (acceptance_sort_log.csv) into the year folder listing,
  for every PDF: doi, accepted_date, status, and any error notes.

Workbook output (OSF headers, monthly tabs + Sheet7):
- Filled columns:
    Filename            : article title (from Crossref via DOI)
    Month               : YYYY-MM-01 based on acceptance month
    Link                : DOI URL
    First author gender : inferred from first author given name
    Last author gender  : inferred from last author given name
    Additional notes    : status flags (e.g., accepted_not_found, doi_not_found, pdf_parse_failed)

USAGE
-----
pip install openpyxl PyPDF2 requests gender-guesser Genderize pandas

python sort_mrm_pdfs_by_acceptance_and_build_workbook.py \
  --year 2025 \
  --pdf-folder "/Users/kahina/Montréal/Cours/GBM6330E_emerging_biotech/CS2_MRI/2025" \
  --popular-names "/Users/kahina/Montréal/Cours/GBM6330E_emerging_biotech/popular_names.csv" \
  --mailto "kahina.baouche@mila.quebec" \
  --use-genderize

Run safe first:
  ... --dry-run
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, Tuple, List

import requests
from openpyxl import Workbook

import PyPDF2
import gender_guesser.detector as gender_detector

# Genderize is optional
try:
    from genderize import Genderize  # type: ignore
    HAVE_GENDERIZE = True
except Exception:
    HAVE_GENDERIZE = False


MONTH_NAMES = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
]

COLUMNS = [
    "Filename",
    "Month",
    "Keywords Matched",
    "Data Availability Statement",
    "False Positive?",
    "Link",
    "Shared code?",
    "Shared data?",
    "Language(s)",
    "Additional notes",
    "First author gender",
    "Last author gender",
]

DOI_RE = re.compile(r"\b10\.\d{4,9}/[^\s<>\"']+\b", re.IGNORECASE)

# Multiple Accepted patterns (after normalization)
# 1) Accepted: 2 April 2025
ACC_DMY = re.compile(r"\bAccepted\b\s*:?\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})\b", re.IGNORECASE)
# 2) Accepted: April 2, 2025
ACC_MDY = re.compile(r"\bAccepted\b\s*:?\s*([A-Za-z]+)\s+(\d{1,2})(?:,)?\s+(\d{4})\b", re.IGNORECASE)
# 3) Accepted: 2025-04-02
ACC_YMD_DASH = re.compile(r"\bAccepted\b\s*:?\s*(\d{4})-(\d{1,2})-(\d{1,2})\b", re.IGNORECASE)
# 4) Accepted: 02/04/2025 or 2/4/2025 (ambiguous; we treat as D/M/Y by default)
ACC_DMY_SLASH = re.compile(r"\bAccepted\b\s*:?\s*(\d{1,2})/(\d{1,2})/(\d{4})\b", re.IGNORECASE)

MONTH_LOOKUP = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


@dataclass
class LogRow:
    pdf: str
    doi: str
    accepted_date: str
    status: str
    note: str


def normalize_text_for_dates(text: str) -> str:
    """
    Normalize PDF-extracted text to improve regex matching.
    - Replace non-breaking spaces with normal spaces
    - Collapse whitespace runs (including newlines) into single spaces
    - Insert spaces between digit/letter boundaries if stuck together (best-effort)
    """
    if not text:
        return ""
    t = text.replace("\u00a0", " ").replace("\t", " ")
    t = re.sub(r"\s+", " ", t)
    # e.g., "Accepted:2April2025" -> "Accepted: 2 April 2025" (best-effort)
    t = re.sub(r"(\d)([A-Za-z])", r"\1 \2", t)
    t = re.sub(r"([A-Za-z])(\d)", r"\1 \2", t)
    return t.strip()


def extract_text_pypdf2(path: Path, max_pages: int = 3) -> str:
    """Extract text from first pages; return '' if PDF can't be parsed."""
    try:
        with path.open("rb") as f:
            reader = PyPDF2.PdfReader(f, strict=False)
            pages = reader.pages
            n = min(max_pages, len(pages))
            chunks: List[str] = []
            for i in range(n):
                try:
                    chunks.append(pages[i].extract_text() or "")
                except Exception:
                    chunks.append("")
            return "\n".join(chunks)
    except Exception as e:
        return ""


def parse_doi(text: str) -> str:
    m = DOI_RE.search(text or "")
    if not m:
        return ""
    return m.group(0).rstrip(").,;]").lower()


def parse_accepted_date(text: str) -> Optional[dt.date]:
    t = normalize_text_for_dates(text)
    if not t:
        return None

    m = ACC_DMY.search(t)
    if m:
        d, mon, y = int(m.group(1)), m.group(2), int(m.group(3))
        mon_i = MONTH_LOOKUP.get(mon.lower(), MONTH_LOOKUP.get(mon.lower()[:3]))
        if mon_i:
            try:
                return dt.date(y, mon_i, d)
            except ValueError:
                return None

    m = ACC_MDY.search(t)
    if m:
        mon, d, y = m.group(1), int(m.group(2)), int(m.group(3))
        mon_i = MONTH_LOOKUP.get(mon.lower(), MONTH_LOOKUP.get(mon.lower()[:3]))
        if mon_i:
            try:
                return dt.date(y, mon_i, d)
            except ValueError:
                return None

    m = ACC_YMD_DASH.search(t)
    if m:
        y, mon, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return dt.date(y, mon, d)
        except ValueError:
            return None

    m = ACC_DMY_SLASH.search(t)
    if m:
        d, mon, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        # Default to D/M/Y (Canada/Europe style). If you want M/D/Y, swap here.
        try:
            return dt.date(y, mon, d)
        except ValueError:
            return None

    return None


def crossref_lookup(doi: str, mailto: str, timeout: int = 30) -> Tuple[str, str, str]:
    """Return (title, first_author_firstname, last_author_firstname) from Crossref."""
    if not doi:
        return "", "", ""
    url = f"https://api.crossref.org/works/{requests.utils.quote(doi)}"
    headers = {"User-Agent": f"mrm-acceptance-sorter (mailto:{mailto})"} if mailto else {}
    try:
        r = requests.get(url, headers=headers, timeout=timeout)
        if not r.ok:
            return "", "", ""
        msg = r.json().get("message", {}) or {}
    except Exception:
        return "", "", ""

    title = ""
    t = msg.get("title")
    if isinstance(t, list) and t:
        title = str(t[0]).strip()
    elif isinstance(t, str):
        title = t.strip()

    authors = msg.get("author") or []
    first_fn = ""
    last_fn = ""
    if isinstance(authors, list) and authors:
        first = authors[0] or {}
        last = authors[-1] or {}
        first_fn = str(first.get("given", "")).strip().split(" ")[0]
        last_fn = str(last.get("given", "")).strip().split(" ")[0]

    return title, first_fn, last_fn


def load_popular_names_csv(path: Optional[Path]) -> Dict[str, str]:
    """
    Supports:
      - wide: columns male,female  (your file)
      - long: name + gender/sex
    """
    if not path:
        return {}
    if not path.exists():
        print(f"WARNING: popular_names.csv not found: {path}", file=sys.stderr)
        return {}

    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return {}
        fields = [c.strip().lower() for c in reader.fieldnames]

        # Wide format: male,female
        if "male" in fields and "female" in fields and len(fields) == 2:
            male_col = reader.fieldnames[fields.index("male")]
            female_col = reader.fieldnames[fields.index("female")]
            out: Dict[str, str] = {}
            for row in reader:
                m = (row.get(male_col) or "").strip()
                if m:
                    out[m.lower()] = "male"
                w = (row.get(female_col) or "").strip()
                if w:
                    out[w.lower()] = "female"
            return out

        # Long format: name + gender/sex
        name_col = None
        gender_col = None
        for cand in ("name", "firstname", "first_name", "first"):
            if cand in fields:
                name_col = reader.fieldnames[fields.index(cand)]
                break
        for cand in ("gender", "sex"):
            if cand in fields:
                gender_col = reader.fieldnames[fields.index(cand)]
                break
        if not name_col or not gender_col:
            print(f"WARNING: Could not detect columns in popular_names.csv. Found: {reader.fieldnames}", file=sys.stderr)
            return {}

        out = {}
        for row in reader:
            nm = (row.get(name_col) or "").strip()
            gd = (row.get(gender_col) or "").strip().lower()
            if not nm:
                continue
            if gd in ("m", "male", "man"):
                out[nm.lower()] = "male"
            elif gd in ("f", "female", "woman"):
                out[nm.lower()] = "female"
            else:
                out[nm.lower()] = "unknown"
        return out


def infer_gender(firstname: str,
                 popular_map: Dict[str, str],
                 det: "gender_detector.Detector",
                 use_genderize: bool = False) -> str:
    fn = (firstname or "").strip().split(" ")[0]
    if not fn:
        return "unknown"
    key = fn.lower()

    if key in popular_map and popular_map[key] in ("male", "female"):
        return popular_map[key]

    gg = det.get_gender(fn)
    if gg in ("male", "female"):
        return gg
    if gg == "mostly_male":
        return "male"
    if gg == "mostly_female":
        return "female"

    if use_genderize and HAVE_GENDERIZE:
        try:
            resp = Genderize().get([fn])
            if resp and isinstance(resp, list):
                g = resp[0].get("gender")
                if g in ("male", "female"):
                    return g
        except Exception:
            pass

    return "unknown"


def create_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for m in MONTH_NAMES:
        ws = wb.create_sheet(m)
        ws.append(COLUMNS)
    ws7 = wb.create_sheet("Sheet7")
    ws7.append(COLUMNS)
    return wb


def month_sheet_name(d: Optional[dt.date]) -> str:
    if not d:
        return "Sheet7"
    return MONTH_NAMES[d.month - 1]


def month_cell_value(d: Optional[dt.date]) -> str:
    if not d:
        return ""
    return f"{d.year:04d}-{d.month:02d}-01"


def ensure_month_folders(base: Path) -> Dict[str, Path]:
    out = {}
    for m in MONTH_NAMES:
        p = base / m
        p.mkdir(parents=True, exist_ok=True)
        out[m] = p
    return out


def write_log_csv(base: Path, rows: List[LogRow]) -> Path:
    out = base / "acceptance_sort_log.csv"
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["pdf", "doi", "accepted_date", "status", "note"])
        for r in rows:
            w.writerow([r.pdf, r.doi, r.accepted_date, r.status, r.note])
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--pdf-folder", required=True)
    ap.add_argument("--popular-names", default="")
    ap.add_argument("--mailto", default="")
    ap.add_argument("--use-genderize", action="store_true")
    ap.add_argument("--max-pages", type=int, default=3)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--move-files", action="store_true",
                    help="Actually move PDFs into month folders. If omitted, files are NOT moved.")
    args = ap.parse_args()

    base = Path(args.pdf_folder).expanduser()
    if not base.exists():
        raise SystemExit(f"--pdf-folder not found: {base}")

    popular_map = load_popular_names_csv(Path(args.popular_names).expanduser() if args.popular_names else None)
    det = gender_detector.Detector(case_sensitive=False)

    month_folders = ensure_month_folders(base)
    wb = create_workbook()

    # Recursive scan so reruns don't miss PDFs already in month folders
    pdfs = sorted([p for p in base.rglob("*.pdf") if p.is_file()])
    if not pdfs:
        print(f"No PDFs found under: {base}")
        return

    log_rows: List[LogRow] = []
    processed = 0

    for pdf in pdfs:
        processed += 1

        text = extract_text_pypdf2(pdf, max_pages=args.max_pages)
        note_parts: List[str] = []

        if not text:
            note_parts.append("pdf_parse_failed")

        doi = parse_doi(text)
        if not doi:
            note_parts.append("doi_not_found")

        acc = parse_accepted_date(text)
        if not acc:
            note_parts.append("accepted_not_found")

        title, first_fn, last_fn = crossref_lookup(doi, mailto=args.mailto) if doi else ("", "", "")
        if not title and doi:
            note_parts.append("crossref_lookup_failed")

        g_first = infer_gender(first_fn, popular_map, det, use_genderize=args.use_genderize)
        g_last = infer_gender(last_fn, popular_map, det, use_genderize=args.use_genderize)

        sheet = month_sheet_name(acc)
        month_val = month_cell_value(acc)
        doi_url = f"https://doi.org/{doi}" if doi else ""

        additional = ";".join(note_parts)

        wb[sheet].append([
            title,           # Filename (title)
            month_val,       # Month (acceptance month)
            "[]",
            "",
            "",
            doi_url,         # Link (DOI URL)
            "",
            "",
            "",
            additional,      # Additional notes
            g_first,
            g_last,
        ])

        status = "ok" if (doi and acc) else "needs_review"
        log_rows.append(LogRow(
            pdf=str(pdf.relative_to(base)),
            doi=doi,
            accepted_date=acc.isoformat() if acc else "",
            status=status,
            note=additional
        ))

        # Move file if requested and accepted found and parse ok
        if args.move_files and acc and "pdf_parse_failed" not in note_parts:
            dest_folder = month_folders[MONTH_NAMES[acc.month - 1]]
            dest_path = dest_folder / pdf.name
            if args.dry_run:
                print(f"[DRY RUN] move: {pdf} -> {dest_path}")
            else:
                if dest_path.exists():
                    dest_path = dest_folder / f"{pdf.stem}__dup{pdf.suffix}"
                # only move if it's not already in the right place
                if pdf.parent.resolve() != dest_folder.resolve():
                    shutil.move(str(pdf), str(dest_path))

    out_xlsx = base / f"{args.year}-MRMH-ReproducibleResearch_acceptance.xlsx"
    out_log = base / "acceptance_sort_log.csv"

    if args.dry_run:
        print(f"[DRY RUN] would write workbook: {out_xlsx}")
        print(f"[DRY RUN] would write log: {out_log}")
    else:
        wb.save(out_xlsx)
        write_log_csv(base, log_rows)
        print(f"Wrote: {out_xlsx}")
        print(f"Wrote: {out_log}")

    print(f"Processed PDFs: {processed}")
    if args.use_genderize and not HAVE_GENDERIZE:
        print("NOTE: --use-genderize set but Genderize is not installed. Install: pip install Genderize")


if __name__ == "__main__":
    main()
