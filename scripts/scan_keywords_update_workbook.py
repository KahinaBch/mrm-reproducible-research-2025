#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scan month folders for PDFs, detect keywords EXACTLY like the original notebook,
and update an existing OSF-style workbook's "Keywords Matched" column.

This script intentionally mirrors the notebook semantics:

- Uses PyPDF2
- For each PDF:
  - For each keyword:
    - For each page:
      - extractText()
      - re.search(keyword, page_text)
      - if match -> record keyword and break out of page loop for that keyword
- Only adds a keyword once per PDF (preserves the keyword list order)

Inputs
------
--year-folder: The year folder containing month subfolders (January..December) and the workbook.
              Example: /Users/.../CS2_MRI/2025

--xlsx: Path to the workbook to update.
        Example: /Users/.../CS2_MRI/2025/2025-MRMH-ReproducibleResearch_acceptance.xlsx

Matching PDFs to workbook rows
------------------------------
We update rows by matching, in this order:

1) If "Link" contains a DOI URL and the PDF text contains that DOI -> match.
2) Else, if workbook "Filename" (article title) is a substring of the PDF filename -> match.
3) Else, if PDF filename is a substring of workbook "Filename" -> match (fallback).

If multiple rows match, we update the first one and log an ambiguity warning.

Outputs
-------
- Updates the workbook in place (makes a timestamped backup by default)
- Writes a CSV log: keyword_scan_log.csv in the year folder

Dependencies
------------
pip install PyPDF2 openpyxl
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import shutil
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import PyPDF2
from openpyxl import load_workbook

MONTH_NAMES = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
]

# NOTE: We keep the same list (and quirks) as the notebook.
# The notebook code has a missing comma between " git " and "osf",
# which concatenates into " git osf". We FIX that here by adding the comma,
# because otherwise "osf" is never searched.
SEARCH_TERMS = [
    "open source",
    "open-source",
    "opensource",
    "open science",
    "github",
    " git ",
    "osf",
    "jupyter",
    "notebook",
    "octave",
    "available online",
    "released",
    "shared",
    " code ",
]

DOI_RE = re.compile(r"\b10\.\d{4,9}/[^\s<>\"']+\b", re.IGNORECASE)


def notebook_style_keyword_scan(pdf_path: Path, search_terms: List[str]) -> List[str]:
    """Replicate the notebook keyword scan behavior (case-sensitive re.search)."""
    found: List[str] = []
    seen = set()
    with pdf_path.open("rb") as f:
        reader = PyPDF2.PdfReader(f, strict=False)
        num_pages = len(reader.pages)

        for keyword in search_terms:
            if keyword in seen:
                continue
            for page_index in range(num_pages):
                page_obj = reader.pages[page_index]
                try:
                    page_text = page_obj.extractText() or ""  # notebook used extractText()
                except Exception:
                    try:
                        page_text = page_obj.extract_text() or ""
                    except Exception:
                        page_text = ""
                if re.search(keyword, page_text) is not None:
                    found.append(keyword)
                    seen.add(keyword)
                    break
    return found


def pdf_contains_doi(pdf_path: Path, doi: str, max_pages: int = 2) -> bool:
    """Best-effort: check whether DOI string appears in first pages of a PDF."""
    if not doi:
        return False
    try:
        with pdf_path.open("rb") as f:
            reader = PyPDF2.PdfReader(f, strict=False)
            n = min(max_pages, len(reader.pages))
            for i in range(n):
                try:
                    t = reader.pages[i].extract_text() or ""
                except Exception:
                    t = ""
                if doi.lower() in t.lower():
                    return True
    except Exception:
        return False
    return False


def parse_doi_from_link(link: str) -> str:
    if not link:
        return ""
    m = DOI_RE.search(link)
    if not m:
        return ""
    return m.group(0).lower()


def build_workbook_index(wb) -> Dict[str, List[Tuple[str,int]]]:
    """
    Build an index for quick access:
      returns dict sheet_name -> list of (doi, row_index) for each data row
    """
    out = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        if not header or "Link" not in header or "Filename" not in header or "Keywords Matched" not in header:
            continue
        link_col = header.index("Link") + 1
        doi_rows = []
        for r in range(2, ws.max_row + 1):
            link = ws.cell(r, link_col).value
            doi = parse_doi_from_link(str(link or ""))
            doi_rows.append((doi, r))
        out[sheet] = doi_rows
    return out


def update_keywords_in_workbook(xlsx_path: Path, year_folder: Path, make_backup: bool = True) -> None:
    wb = load_workbook(xlsx_path)
    index = build_workbook_index(wb)

    log_rows = []
    updated = 0
    scanned = 0

    for month in MONTH_NAMES:
        month_dir = year_folder / month
        if not month_dir.exists():
            continue

        pdfs = sorted([p for p in month_dir.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"])
        for pdf in pdfs:
            scanned += 1

            # Scan keywords notebook-style
            try:
                kws = notebook_style_keyword_scan(pdf, SEARCH_TERMS)
            except Exception as e:
                log_rows.append([month, pdf.name, "", "scan_failed", str(e)])
                continue

            if not kws:
                log_rows.append([month, pdf.name, "", "no_keywords", ""])
                continue

            # Try to match workbook row using DOI-in-link first
            matched_sheet = None
            matched_row = None

            # Extract DOI from PDF quickly (best-effort from first pages)
            doi_in_pdf = ""
            try:
                with pdf.open("rb") as f:
                    reader = PyPDF2.PdfReader(f, strict=False)
                    # look in first 2 pages
                    for i in range(min(2, len(reader.pages))):
                        t = ""
                        try:
                            t = reader.pages[i].extract_text() or ""
                        except Exception:
                            pass
                        m = DOI_RE.search(t)
                        if m:
                            doi_in_pdf = m.group(0).lower()
                            break
            except Exception:
                pass

            if doi_in_pdf:
                # Search row with same doi
                for sheet, rows in index.items():
                    for doi, r in rows:
                        if doi and doi == doi_in_pdf:
                            matched_sheet, matched_row = sheet, r
                            break
                    if matched_sheet:
                        break

            # If no DOI match, attempt filename/title heuristic match within the same month sheet first
            if not matched_sheet:
                ws = wb[month] if month in wb.sheetnames else None
                if ws:
                    header = [c.value for c in ws[1]]
                    fn_col = header.index("Filename") + 1
                    for r in range(2, ws.max_row + 1):
                        title = str(ws.cell(r, fn_col).value or "")
                        if title and (title.lower() in pdf.name.lower() or pdf.stem.lower() in title.lower()):
                            matched_sheet, matched_row = month, r
                            break

            if not matched_sheet:
                log_rows.append([month, pdf.name, str(kws), "no_match_in_xlsx", doi_in_pdf])
                continue

            ws = wb[matched_sheet]
            header = [c.value for c in ws[1]]
            kw_col = header.index("Keywords Matched") + 1
            ws.cell(matched_row, kw_col).value = str(kws)

            updated += 1
            log_rows.append([month, pdf.name, str(kws), f"updated:{matched_sheet}!{matched_row}", doi_in_pdf])

    if make_backup:
        backup = xlsx_path.with_suffix(f".backup_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(xlsx_path, backup)

    wb.save(xlsx_path)

    # write log
    log_path = year_folder / "keyword_scan_log.csv"
    with log_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["month_folder", "pdf", "keywords", "status", "doi_found_in_pdf"])
        for row in log_rows:
            w.writerow(row)

    print(f"Scanned PDFs: {scanned}")
    print(f"Updated rows: {updated}")
    print(f"Workbook updated: {xlsx_path}")
    print(f"Log written: {log_path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year-folder", required=True, help="Folder with month subfolders (January..December)")
    ap.add_argument("--xlsx", required=True, help="Workbook to update (OSF format)")
    ap.add_argument("--no-backup", action="store_true", help="Do not create a backup XLSX")
    args = ap.parse_args()

    year_folder = Path(args.year_folder).expanduser()
    xlsx_path = Path(args.xlsx).expanduser()

    update_keywords_in_workbook(
        xlsx_path=xlsx_path,
        year_folder=year_folder,
        make_backup=(not args.no_backup),
    )


if __name__ == "__main__":
    main()
